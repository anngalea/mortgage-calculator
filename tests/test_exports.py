from io import BytesIO

from openpyxl import load_workbook

from exports import excel_export, ledger_csv, monthly_summary_csv
from mortgage_engine import project_mortgage
from sample_data import DEFAULT_MORTGAGE, DEFAULT_RATE_SCHEDULE, DEFAULT_SETTINGS


def test_exports_are_non_empty_and_excel_has_required_sheets():
    result = project_mortgage(DEFAULT_MORTGAGE, DEFAULT_RATE_SCHEDULE, [], DEFAULT_SETTINGS)

    workbook_bytes = excel_export(result, DEFAULT_MORTGAGE, DEFAULT_RATE_SCHEDULE, [])
    workbook = load_workbook(BytesIO(workbook_bytes))

    assert "Dashboard Summary" in workbook.sheetnames
    assert "Transaction Ledger" in workbook.sheetnames
    assert "Monthly Summary" in workbook.sheetnames
    assert ledger_csv(result)
    assert monthly_summary_csv(result)
